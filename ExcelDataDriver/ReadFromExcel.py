import openpyxl
path = "Your excel file location"
work_book = openpyxl.load_workbook(path)
sheet = work_book.active      #work_book.get_sheet_by_name("Sheet1")
max_rows = sheet.max_row
max_coloumns = sheet.max_column
print(max_rows)
print(max_coloumns)
for r in range(1,max_rows+1):
    for c in range(1,max_coloumns+1):
        print(sheet.cell(row = r,column = c).value,end= "       ")
    print()
    print()
    print()
    print()
    print()
    print()
    print()



